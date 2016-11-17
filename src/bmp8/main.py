#!/usr/bin/python
# -*- coding: utf-8 -*-

#
#  This file is part of the `bmp8` Python module
#
#  Copyright (c) 2016 - EMBL-EBI
#
#  File author(s): Dénes Türei (turei.denes@gmail.com)
#
#  Distributed under the GPLv3 License.
#  See accompanying file LICENSE.txt or copy at
#      http://www.gnu.org/licenses/gpl-3.0.html
#
#  Website: http://www.ebi.ac.uk/~denes
#

from future.utils import iteritems
from past.builtins import xrange, range, reduce

import os
import sys
import imp
import re
import itertools
import copy

import xlrd
import openpyxl

import numpy as np

import pypath

if 'unicode' not in __builtins__:
    unicode = str

class Bmp8(object):
    
    def __init__(self,
                 dirBase = '..',
                 fnIdMapping = 'PEX100_Layout.csv',
                 fnXlsFullData = 'AssayData-Cambridge-Peirce-150515.xlsx',
                 fnTable = 'bmp8.csv',
                 ncbi_tax_id = 10090,
                 org_strict = True,
                 flex_resnum = True):
        
        self.dirBase = dirBase
        self.fnXlsFullData = os.path.join(
            self.dirBase,
            fnXlsFullData
        )
        self.fnIdMapping = os.path.join(
            self.dirBase,
            fnIdMapping
        )
        self.fnTable = os.path.join(
            self.dirBase,
            fnTable
        )
        self.reAnnot = re.compile(r'([\-\s/\.,\(\)\+A-Za-z0-9]{2,}) '\
            r'\(([A-Z][a-z]+)-?([A-Za-z0-9/]*)\)')
        self.reRes = re.compile(r'([A-Z]?[a-z]*)([0-9]+)')
        self.dAaletters = {
            'Thr': 'T',
            'Tyr': 'Y',
            'Ser': 'S'
        }
        self.dOrg = {
            'H': 9606,
            'M': 10090,
            'R': 10116
        }
        
        self.ncbi_tax_id = ncbi_tax_id
        self.org_strict = org_strict
        self.flex_resnum = flex_resnum
        
        self.standards = ['Beta actin', 'GAPDH', 'PKC pan activation site']
        self.dStd = {}
        
        self.dMissingHomologs = {
            'IKK-alpha': 'Q60680',
            'IKK-alpha/beta': 'Q60680',
            'GAPDH': 'P16858',
            'Calmodulin': 'P62204',
            'TYK2': 'Q9R117',
            'Beta actin': 'P60710',
            'Met': 'P16056',
            'ATP1A1/Na+K+ ATPase1': 'Q8VDN2',
            'GRK1': 'Q9WVL4',
            'MAP3K7/TAK1': 'Q62073',
            'Rel': 'P15307',
            'Gab2': 'Q9Z1S8',
            'DDX5/DEAD-box protein 5': 'Q61656',
            'MARCKS': 'P26645',
            'PLD1': 'Q9Z280',
            'CD45': 'P06800',
            'Pim-1': 'P06803',
            'MAP3K1/MEKK1': 'P53349',
            'Progesterone Receptor': 'Q00175',
            'PAK1': 'O88643',
            'RapGEF1': 'Q8C5V7',
            'PAK1/2': ['O88643', 'Q8CIN4'],
            'PAK1/2/3': ['O88643', 'Q8CIN4', 'Q61036'],
            'CD227/mucin 1': 'Q02496',
            'HDAC3': 'O88895',
            'TSC2': 'Q61037',
            'Tuberin/TSC2': 'Q61037',
            'CaMK4': 'P08414',
            'STAT2': 'Q9WVL2',
            'AurB/C': ['O70126', 'O88445'],
            'JAK1': 'P52332',
            'NFAT4': 'P97305',
            'HDAC5': 'Q9Z2V6',
            'VEGFR2': 'P35918',
            'STAT1': 'P42225',
            'EPHA2/3/4': ['Q03145', 'Q03137', 'P29319'],
            'Histone H3.1': ['P68433', 'P84228', 'P84244']
        }
        
        self.dAssayRefCorrection = {
            ('Kv1.3/KCNA3', 135): {
                9606: ('Y', 187),
                10090: ('Y', 140),
                10116: ('Y', 137)
            },
            ('PKC delta', 505): {
                9606: ('T', 507)
            },
            ('Opioid Receptor', 375): {
                9606: ('S', 377)
            },
            ('IkB-epsilon', 22): {
                9606: ('S', 161)
            },
            ('HSL', 563): {
                9606: ('S', 853),
                10090: ('S', 557),
                10116: ('S', 863)
            },
            ('KSR', 392): {
                9606: ('S', 406),
                10116: ('S', 40)
            },
            ('BAD', 155): {
                9606: ('S', 118),
                10116: ('S', 156)
            },
            ('BAD', 112): {
                9606: ('S', 75),
                10116: ('S', 113)
            },
            # this EPHB2
            ('EPHB1/2', 604): {
                9606: ('Y', 596),
                10090: ('Y', 596),
                10116: ('Y', 709)
            },
            ('NMDAR2B', 1472): {
                9606: ('Y', 1474)
            },
            ('SRF', 99): {
                9606: ('S', 103)
            },
            ('JAK1', 1022): {
                9606: ('Y', 1034),
                10090: ('Y', 1033),
                10116: ('Y', 1078)
            },
            ('IkB-beta', 19): {
                9606: ('S', 19)
            },
            ('Myosin regulatory light chain 2', 18): {
                9606: ('S', 20),
                10090: ('S', 20),
                10116: ('S', 20)
            },
            ('BAD', 128): {
                9606: ('S', 91),
                10116: ('S', 129)
            },
            ('BAD', 136): {
                9606: ('S', 99),
                10116: ('S', 137)
            },
            ('RelB', 552): {
                9606: ('S', 573)
            },
            ('IL3RB', 593): {
                10090: ('Y', 595)
            },
            ('BIM', 65): {
                9606: ('S', 69)
            },
            ('CaMK4', 196): {
                9606: ('T', 200)
            }
        }
    
    def reload(self):
        modname = self.__class__.__module__
        mod = __import__(modname, fromlist = [modname.split('.')[0]])
        imp.reload(mod)
        new = getattr(mod, self.__class__.__name__)
        setattr(self, '__class__', new)
    
    #
    # Loading experimental data.
    #
    
    def init(self):
        self.init_pypath()
        self.idmapping()
        self.read_tables()
        self.load_ptms()
        self.load_seq()
        self.create_ptms()
        self.ptms_lookup()
    
    def network(self, extra_proteins = [], edges_percentile = 50, pfile = None):
        self.collect_proteins()
        self.load_network(pfile = pfile)
        self.get_network(keep_also = extra_proteins)
        self.sparsen_network(perc = edges_percentile)
    
    def export_tables(self):
        pass
    
    def init_pypath(self):
        """
        Initializes a PyPath object for mapping and network building.
        """
        self.pa = pypath.PyPath(ncbi_tax_id = self.ncbi_tax_id)
    
    def read_tables(self):
        """
        Reads all data tables.
        """
        self.read_signals()
        self.read_coeffvar()
        self.read_normalized()
        self.read_psites_diffs()
    
    def idmapping(self):
        """
        Builds the idmapping table in order to translate the custom names
        to UniProts already at loading the table.
        """
        self.read_idmapping()
        self.read_organism_specificities()
        if self.ncbi_tax_id == 10090:
            self.mousedict()
            self.idmapping2mouse()
            self.dNamesUniprots = self.dNamesMouseUniprots
        else:
            self.dNamesHumanUniprots = \
                dict(
                    map(
                        lambda x:
                            (
                                x[0],
                                x[1][0]
                            ),
                        iteritems(self.dNamesIds)
                    )
                )
            
            self.dNamesUniprots = self.dNamesHumanUniprots
        
        self.dOrgSpecUniprot = \
            dict(
                itertools.chain(
                    *map(
                        lambda i:
                            map(
                                lambda u:
                                    (
                                        (u, i[0][1], i[0][2]),
                                        i[1]
                                    ),
                                self.dNamesUniprots[i[0][0]]
                            ),
                        iteritems(self.dOrgSpec)
                    )
                )
            )
    
    def read_psites_diffs(self):
        """
        Reads the measured values for each phosphosite at each condition.
        """
        
        self.read_data('Psite', 16, 9, 7, 582, 3)
    
    def read_signals(self):
        """
        Reads the mean of duplicates signal values.
        """
        
        self.read_data('Signal', 0, 9, 5, 1320, 1)
    
    def read_coeffvar(self):
        """
        Reads coefficient of variation values for each pair of spots.
        """
        
        self.read_data('CVar', 0, 9, 10, 1320, 6)
    
    def read_normalized(self):
        """
        Reads the values normalized to median.
        """
        
        self.read_data('Norm', 0, 9, 15, 1320, 11)
    
    def read_data(self, name, col, row, ncol, nrow, vcol):
        """
        Reads a data table with its annotations.
        """
        
        if not hasattr(self, 'llFullData') or self.llFullData is None:
            self.read_raw()
        
        name = name.capitalize()
        
        attr = 'll%s' % name
        
        setattr(self, attr,
                self.ll_table_slice(self.llFullData, col, row, ncol, nrow))
        
        annot, data = self.get_arrays(getattr(self, attr), vcol)
        
        setattr(self, 'a%sAnnot' % name, annot)
        setattr(self, 'a%sData' % name, data)
        
        self.dStd[name] = self.get_standards(name, vcol)
    
    def get_standards(self, name, vcol):
        """
        Gets the values from the standards.
        """
        return \
            dict(
                map(
                    lambda l:
                        (
                            l[0],
                            list(map(float, l[vcol:]))
                        ),
                    filter(
                        lambda l:
                            l[0] in self.standards,
                        getattr(self, 'll%s' % name)
                    )
                )
            )
    
    def read_raw(self):
        """
        Reads the full xls table.
        """
        self.llFullData = self.read_xls(self.fnXlsFullData)
    
    def read_xls(self, xls_file, sheet = 0, csv_file = None,
        return_table = True):
        """
        Generic function to read MS Excel XLS file. Converts one sheet
        to CSV, or returns as a list of lists
        """
        table = []
        try:
            book = xlrd.open_workbook(xls_file, on_demand = True)
            try:
                if type(sheet) is int:
                    sheet = book.sheet_by_index(sheet)
                else:
                    sheet = book.sheet_by_name(sheet)
            except xlrd.biffh.XLRDError:
                sheet = book.sheet_by_index(0)
            table = [[unicode(c.value) \
                for c in sheet.row(i)] \
                for i in xrange(sheet.nrows)]
        except IOError:
            sys.stdout.write('No such file: %s\n' % xls_file)
            sys.stdout.flush()
        except:
            try:
                book = openpyxl.load_workbook(filename = xls_file,
                    read_only = True)
            except:
                sys.stdout.write('\tCould not open xls: %s\n' % xls_file)
                if not os.path.exists(xls_file):
                    sys.stdout.write('\tFile does not exist.\n')
                sys.stdout.flush()
            try:
                if type(sheet) is int:
                    sheet = book.worksheets[sheet]
                else:
                    sheet = book[sheet]
            except:
                sheet = book.worksheets[0]
            cells = sheet.get_squared_range(1, 1,
                sheet.max_column, sheet.max_row)
            table = \
                list(
                    map(lambda row:
                        list(
                            map(lambda c:
                                unicode(c.value),
                                row
                            )
                        ),
                        cells
                    )
                )
        if csv_file:
            with open(csv_file, 'w') as csv:
                csv.write('\n'.join(['\t'.join(r) for r in table]))
        if not return_table:
            table = None
        if 'book' in locals() and hasattr(book, 'release_resources'):
            book.release_resources()
        return table
    
    def ll_table_slice(self, table, col, row, ncol = None, nrow = None):
        """
        Returns a slice from a list-of-lists table, starting at (row, col),
        height and width of (nrow, ncol).
        """
        cend = None if ncol is None else col + ncol
        rend = None if nrow is None else row + nrow
        return list(map(lambda l: list(l)[col:cend], table[row:rend]))
    
    def get_arrays(self, llTable, vcol):
        """
        Returns annotations and data as numpy arrays.
        Requires list-of-lists table and
        the column number where data values start.
        """
        
        return \
            list(
                map(
                    lambda a:
                        np.array(a[0], dtype = a[1]),
                    # zip data and dtype
                    zip(
                        # zip annotations and values
                        zip(
                            *list(
                                # filter organism mismatches
                                filter(
                                    lambda r:
                                        not self.org_strict or \
                                            self.ncbi_tax_id in \
                                            self.dOrgSpec[(r[0][1], r[0][3], r[0][4])],
                                    # chain rows & residues
                                    itertools.chain(
                                        *map(
                                            # iterate rows
                                            lambda ll:
                                                list(
                                                    itertools.chain(
                                                        *map(
                                                            # iterate residues
                                                            lambda res:
                                                                map(
                                                                    lambda u:
                                                                        (
                                                                            # annotation
                                                                            [
                                                                                u,
                                                                                ll[0][0],
                                                                                ll[0][1],
                                                                                res[0],
                                                                                res[1]
                                                                            ],
                                                                            # values
                                                                            ll[1]
                                                                        ),
                                                                    self.dNamesUniprots[ll[0][0]]
                                                                ),
                                                            self.get_residues(ll[0][2])
                                                        )
                                                    )
                                                ),
                                            map(
                                                # for each row, match annotation
                                                # and convert data to float
                                                lambda l:
                                                    (
                                                        self.reAnnot.\
                                                            match(l[0]).groups(0),
                                                        list(map(float, l[vcol:]))
                                                    ),
                                                # starting from list of lists
                                                filter(
                                                    # filter standards
                                                    lambda l:
                                                        l[0] not in self.standards,
                                                    llTable
                                                )
                                            )
                                        )
                                    )
                                )
                            )
                        ),
                        # dtypes for annot. and data
                        [np.object, np.float]
                    )
                )
            )
    
    def get_residues(self, res):
        """
        Matches residues in string,
        returns tuples of tuples with residue names and numbers.
        """
        ress = self.reRes.findall(res)
        aa = ''
        sres = []
        for r in ress:
            if r[0] in self.dAaletters:
                aa = self.dAaletters[r[0]]
            sres.append((aa, int(r[1])))
        return tuple(sres)
    
    def read_organism_specificities(self):
        """
        Reads the organism specificities of the assay probes.
        These are either for human, mouse, rat, or any combination of these.
        """
        with open(self.fnIdMapping, 'r') as f:
            self.dOrgSpec = \
                dict(
                    itertools.chain(
                        *map(
                            lambda p:
                                map(
                                    lambda res:
                                        (
                                            (p[0][0], res[0], res[1]),
                                            list(
                                                map(
                                                    lambda s:
                                                        self.dOrg[s],
                                                    p[1]
                                                )
                                            )
                                        ),
                                    self.get_residues(p[0][2])
                                ),
                            map(
                                lambda l:
                                    (
                                        self.reAnnot.match(l[1]).groups(0),
                                        l[2]
                                    ),
                                filter(
                                    lambda l:
                                        l[1].strip() not in self.standards,
                                    map(
                                        lambda l:
                                            l.split('\t'),
                                        f
                                    )
                                )
                            )
                        )
                    )
                )
    
    def read_idmapping(self):
        """
        Reads the mapping between the assay custom names and
        standard IDs (UniProt and Entrez Gene).
        """
        with open(self.fnIdMapping, 'r') as f:
            self.dNamesIds = \
                dict(
                    map(
                        lambda l:
                            (
                                self.reAnnot.match(l[1]).groups(0)[0] \
                                    if l[1] not in self.standards \
                                    else l[1], # name
                                (
                                    list(
                                        map(
                                            lambda u:
                                                u.strip(),
                                            l[3].strip().split('/'), # uniprot
                                        )
                                    ),
                                    list(
                                        map(
                                            lambda e:
                                                e.strip(),
                                            l[4].strip().replace(
                                                '\xa0', '').split('/')# entrez
                                        )
                                    )
                                )
                            ),
                        map(
                            lambda l:
                                l.split('\t'),
                            f
                        )
                    )
                )
    
    def idmapping2mouse(self):
        """
        Creates a dict of names to mouse UniProt IDs.
        """
        self.dNamesMouseUniprots = \
            dict(
                map(
                    lambda i:
                        (
                            i[0],
                            list(
                                itertools.chain(
                                    *map(
                                        lambda e:
                                            self.pa.mapper.map_name(e,
                                                                    'entrez',
                                                                    'uniprot',
                                                                    10090),
                                        reduce(
                                            lambda s1, s2:
                                                s1 | s2,
                                            map(
                                                lambda h:
                                                    self.dHuman2Mouse[h],
                                                filter(
                                                    lambda h:
                                                        h in self.dHuman2Mouse,
                                                    i[1][1]
                                                )
                                            ),
                                            set([])
                                        )
                                    )
                                )
                            )
                        ),
                    iteritems(self.dNamesIds)
                )
            )
        
        # adding missing homologs
        for n in self.dNamesMouseUniprots.keys():
            if n in self.dMissingHomologs:
                m = self.dMissingHomologs[n]
                if type(m) is list:
                    self.dNamesMouseUniprots[n].extend(m)
                else:
                    self.dNamesMouseUniprots[n].append(m)
                self.dNamesMouseUniprots[n] = \
                    list(set(self.dNamesMouseUniprots[n]))
    
    def mousedict(self):
        """
        Reads human to mouse mapping from HomoloGene.
        """
        self.dHuman2Mouse = pypath.dataio.homologene_dict(9606, 10090, 'entrez')
    
    #
    # Loading database data.
    #
    
    def load_ptms(self):
        """
        Obtains a list of all kinase-substrate interactions from
        Signor, phosphoELM, PhosphoSite and dbPTM.
        Creates a dict `dPhosDb` where (kinase, substrate) UniProts are keys
        and list of PTMs are values.
        """
        self.dPhosDb = {}
        
        if self.ncbi_tax_id == 10090:
            lKinSub = itertools.chain(
                self.pa.load_signor_ptms(return_raw=True),
                self.pa.load_phosphoelm(return_raw=True),
                self.pa.load_dbptm(return_raw=True),
                self.pa.load_psite_phos(return_raw=True)
            )
        else:
            lKinSub = itertools.chain(
                self.pa.load_signor_ptms(return_raw=True),
                self.pa.load_li2012_ptms(return_raw=True),
                self.pa.load_hprd_ptms(return_raw=True),
                self.pa.load_mimp_dmi(return_raw=True),
                self.pa.load_pnetworks_dmi(return_raw=True),
                self.pa.load_phosphoelm(return_raw=True),
                self.pa.load_dbptm(return_raw=True),
                self.pa.load_psite_phos(return_raw=True)
            )
        
        for ksub in lKinSub:
            if ksub.ptm.protein not in self.dPhosDb:
                self.dPhosDb[ksub.ptm.protein] = []
            self.dPhosDb[ksub.ptm.protein].append(ksub)
    
    def load_seq(self):
        """
        Loads protein sequences from UniProt.
        """
        self.dSeq = pypath.uniprot_input.swissprot_seq(self.ncbi_tax_id,
                                                       isoforms = True)
    
    def create_ptms(self):
        """
        Creates PTM objects for each PTM on array.
        """
        self.dAssaySub = {}
        self.setSeqMismatch = set([])
        self.setSeqMismatchName = set(map(lambda p: (p[1], p[3], p[4]),
                                          self.aPsiteAnnot))
        self.setSeqMismatchOrg = set([])
        self.setSeqMismatchNameOrg = set(
                                      filter(
                                          lambda k:
                                              self.ncbi_tax_id in self.dOrgSpec[k],
                                            map(lambda p: (p[1], p[3], p[4]),
                                                self.aPsiteAnnot)
                                        ))
        
        self.dResUpdate = {}
        
        for psite in self.aPsiteAnnot:
            
            key = (psite[0], psite[3], psite[4])
            nkey = (psite[1], psite[3], psite[4])
            
            # 
            self.setSeqMismatch.add(key)
            
            if psite[0] in self.dSeq:
                
                for isof in self.dSeq[psite[0]].isoforms():
                    
                    resnum = psite[4]
                    resaa = psite[3]
                    
                    if (psite[1], resnum) in self.dAssayRefCorrection and \
                        self.ncbi_tax_id in self.dAssayRefCorrection[
                        (psite[1], resnum)]:
                        
                        resaa, resnum = \
                            self.dAssayRefCorrection[(psite[1], resnum)][
                                self.ncbi_tax_id]
                    
                    if not self.dSeq[psite[0]].match(resaa,
                                                     resnum, isoform = isof):
                        resnum = psite[4] + 1
                        if not self.flex_resnum or \
                            not self.dSeq[psite[0]].match(resaa,
                                                     resnum, isoform = isof):
                            continue
                        
                    if self.dSeq[psite[0]].match(resaa,
                                                     resnum, isoform = isof):
                        
                        # updating the residue number in array
                        if resnum != psite[4]:
                            self.dResUpdate[(psite[0], psite[4])] = \
                                (psite[0], resnum)
                            psite[4] = resnum
                        
                        res = pypath.intera.Residue(resnum,
                                                    resaa,
                                                    psite[0],
                                                    isoform = isof)
                        reg = self.dSeq[psite[0]].get_region(
                            residue = resnum, flanking = 7, isoform = isof)
                        mot = pypath.intera.Motif(psite[0], reg[0], reg[1],
                                                  instance = reg[2],
                                                  isoform = isof)
                        ptm = pypath.intera.Ptm(psite[0],
                                                residue = res,
                                                motif = mot,
                                                typ = 'phosphorylation',
                                                isoform = isof)
                        
                        if psite[0] not in self.dAssaySub:
                            self.dAssaySub[psite[0]] = []
                        self.dAssaySub[psite[0]].append(ptm)
                        
                        # this matched once, so removed to
                        self.setSeqMismatch.remove(key)
                        if nkey in self.setSeqMismatchName:
                            self.setSeqMismatchName.remove(nkey)
                        if nkey in self.setSeqMismatchNameOrg:
                            self.setSeqMismatchNameOrg.remove(nkey)
                        
                        break # process only the first matching isoform
            
            else:
                sys.stdout.write('\t!! No sequence available for %s!\n' % \
                    psite[0])
            
            if key in self.setSeqMismatch:
                
                # here the organism matches, but the sequence mismatches:
                if self.ncbi_tax_id in self.dOrgSpec[nkey]:
                    self.setSeqMismatchOrg.add(key)
            
        # updating residue numbers in all arrays:
        # arrays
        for a in [self.aSignalAnnot, self.aCvarAnnot, self.aNormAnnot]:
            # rows
            for r in a:
                key = (r[0], r[4])
                if key in self.dResUpdate:
                    r[4] = self.dResUpdate[key][1]
        
        for k in list(self.dOrgSpecUniprot.keys()):
            key = (k[0], k[2])
            if key in self.dResUpdate:
                self.dOrgSpecUniprot[(k[0], k[1], self.dResUpdate[key][1])] = \
                    self.dOrgSpecUniprot[k]
                del self.dOrgSpecUniprot[k]
    
    def export_table(self):
        """
        Export a table with number of kinases for each substrate PTM.
        """
        
        def get_pratio(a, ckey, pkey, lnums, cnum):
            return a[lnums[ckey],cnum] / a[lnums[pkey],cnum]
        
        self.lDataCols = [('Control', 0), ('BMP8b', 1), ('NE', 2), ('BMP8b_NE', 3)]
        self.lTableHdr = ['uniprot', 'gsymbol', 'name', 'numof_kin', 'degree',
                          'resaa', 'resnum', 'group',
                          'std_gapdh', 'std_actin', 'std_pkc',
                          'signal', 'ctrl_signal',
                          'cv', 'ctrl_cv',
                          'norm',
                          'norm_actin',
                          'norm_gapdh',
                          'norm_pkc',
                          'ctrl_norm',
                          'ctrl_norm_actin',
                          'ctrl_norm_gapdh',
                          'ctrl_norm_pkc',
                          'phos',
                          'pratio', 'ctrl_pratio',
                          'pratio_actin', 'ctrl_pratio_actin',
                          'pratio_gapdh', 'ctrl_pratio_gapdh',
                          'pratio_pkc', 'ctrl_pratio_pkc',
                          'fc', 'fc_actin', 'fc_gapdh', 'fc_pkc']
        llTable = []
        dDataLnum = dict(map(lambda i: ((i[1][0], i[1][2], i[1][3], i[1][4]), i[0]), enumerate(self.aSignalAnnot)))
        self.dDataLnum = dDataLnum
        
        std_actin = np.array(self.dStd['Signal']['Beta actin'])
        std_actin = std_actin / std_actin.min()
        std_gapdh = np.array(self.dStd['Signal']['GAPDH'])
        std_gapdh = std_gapdh / std_gapdh.min()
        std_pkc = np.array(self.dStd['Signal']['PKC pan activation site'])
        std_pkc = std_pkc / std_pkc.min()
        
        # normalized to standard
        self.aSignalActin = self.aSignalData / std_actin
        self.aSignalGapdh = self.aSignalData / std_gapdh
        self.aSignalPkc   = self.aSignalData / std_pkc
        
        # normalized to median
        self.aSignalActinNorm = self.aSignalActin / np.median(self.aSignalActin, axis = 0)
        self.aSignalGapdhNorm = self.aSignalGapdh / np.median(self.aSignalGapdh, axis = 0)
        self.aSignalPkcNorm   = self.aSignalPkc   / np.median(self.aSignalPkc, axis = 0)
        
        for i, annot in enumerate(self.aPsiteAnnot):
            uniprot = annot[0]
            gss = self.pa.mapper.map_name(uniprot, 'uniprot', 'genesymbol', self.ncbi_tax_id)
            genesymbol = gss[0] if len(gss) else uniprot
            key = (uniprot, annot[3], annot[4])
            pkey = (uniprot, 'Phospho', annot[3], annot[4])
            ckey = (uniprot, 'Ab', '', annot[4])
            degree = self.pa.graph.vs.select(name = uniprot)[0].degree() \
                if uniprot in self.pa.graph.vs['name'] else 0
            if key not in self.dKinNum:
                continue
            
            ctrl_pratio = get_pratio(self.aSignalData, ckey, pkey, dDataLnum, 0)
            
            ctrl_pratio_actin = get_pratio(self.aSignalActinNorm, ckey,
                                           pkey, dDataLnum, 0)
            ctrl_pratio_gapdh = get_pratio(self.aSignalGapdhNorm, ckey,
                                           pkey, dDataLnum, 0)
            ctrl_pratio_pkc   = get_pratio(self.aSignalPkcNorm,   ckey,
                                           pkey, dDataLnum, 0)
            
            for group, cnum in self.lDataCols:
                pratio = get_pratio(self.aNormData, ckey, pkey, dDataLnum, cnum)
                
                pratio_actin = get_pratio(self.aSignalActinNorm, ckey, pkey,
                                          dDataLnum, cnum)
                pratio_gapdh = get_pratio(self.aSignalGapdhNorm, ckey, pkey,
                                          dDataLnum, cnum)
                pratio_pkc   = get_pratio(self.aSignalPkcNorm,   ckey, pkey,
                                          dDataLnum, cnum)
                
                for dkey, phos in [(ckey, 'np'), (pkey, 'p')]:
                    
                    llTable.append(
                        [
                            uniprot,
                            genesymbol,
                            annot[1],
                            '%u' % self.dKinNum[key],
                            '%u' % degree,
                            annot[3],
                            '%u' % annot[4],
                            group,
                            '%.04f' % self.dStd['Signal'][
                                'GAPDH'][cnum], # GAPDH standard
                            '%.04f' % self.dStd['Signal'][
                                'Beta actin'][cnum], # Actin standard
                            '%.04f' % self.dStd['Signal'][
                                'PKC pan activation site'][cnum], # PKC standard
                            '%.04f' % self.aSignalData[dDataLnum[dkey],cnum], # signal
                            '%.04f' % self.aSignalData[dDataLnum[dkey],0],
                            '%.04f' % self.aCvarData[dDataLnum[dkey],cnum], # CV
                            '%.04f' % self.aCvarData[dDataLnum[dkey],0],
                            '%.04f' % self.aNormData[dDataLnum[dkey],cnum], # norm
                            '%.04f' % self.aSignalActinNorm[dDataLnum[dkey],cnum],
                            '%.04f' % self.aSignalGapdhNorm[dDataLnum[dkey],cnum],
                            '%.04f' % self.aSignalPkcNorm[dDataLnum[dkey],cnum],
                            '%.04f' % self.aNormData[dDataLnum[dkey],0],
                            '%.04f' % self.aSignalActinNorm[dDataLnum[dkey],0],
                            '%.04f' % self.aSignalGapdhNorm[dDataLnum[dkey],0],
                            '%.04f' % self.aSignalPkcNorm[dDataLnum[dkey],0],
                            phos,
                            '%.04f' % pratio, # pratio
                            '%.04f' % ctrl_pratio,
                            '%.04f' % pratio_actin,
                            '%.04f' % ctrl_pratio_actin,
                            '%.04f' % pratio_gapdh,
                            '%.04f' % ctrl_pratio_gapdh,
                            '%.04f' % pratio_pkc,
                            '%.04f' % ctrl_pratio_pkc,
                            '%.04f' % (pratio / ctrl_pratio), # fold change
                            '%.04f' % (ctrl_pratio_actin / pratio_actin),
                            '%.04f' % (ctrl_pratio_gapdh / pratio_gapdh),
                            '%.04f' % (ctrl_pratio_pkc / pratio_pkc)
                            
                        ]
                    )
        
        self.llTable = llTable
        self.sTable = '%s\n' % '\t'.join(self.lTableHdr)
        self.sTable = '%s%s' % (self.sTable, '\n'.join(map(lambda l: '%s' % '\t'.join(l), self.llTable)))
        
        with open(self.fnTable, 'w') as f:
            f.write(self.sTable)
    
    def ptms_lookup(self):
        """
        Looks up the PTMs on array in the databases data.
        Creates list with kinase-substrate interactions
        targeting the substrates on array.
        """
        self.dKinAssaySub = {}
        self.dKinNum = {}
        
        for uniprot, ptms in iteritems(self.dAssaySub):
            
            if uniprot not in self.dKinAssaySub:
                self.dKinAssaySub[uniprot] = []
            
            kss = self.dPhosDb[uniprot] if uniprot in self.dPhosDb else []
            
            for ptm in ptms:
                
                nKin = 0
                for ks in kss:
                    if ptm in ks:
                        self.dKinAssaySub[uniprot].append(ks)
                        nKin += 1
                
                self.dKinNum[(uniprot,
                                ptm.residue.name,
                                ptm.residue.number)] = nKin
        
        sys.stdout.write('\t:: Found %u kinase-substrate interactions for\n'\
            '\t   %s phosphorylation sites.\n'\
            '\t   %.02f kinase for one site in average.\n'\
            '\t   For %u sites no kinase found.\n'\
            '\t   Additional %u sites could not be found in UniProt sequences.\n'\
            '\t   The total number of sites is %u.\n'\
            '\t   %u of these sites should be valid for this organism, and\n'\
            '\t   %u mismatches despite should match for this organism.\n'\
            '\t   Excluding redundant combinations, overall %u sites mismatches.\n' % (
                sum(map(len, self.dKinAssaySub.values())),
                len(list(filter(bool, self.dKinNum.values()))),
                np.mean(list(self.dKinNum.values())),
                len(list(filter(lambda n: n == 0, self.dKinNum.values()))),
                len(self.setSeqMismatch),
                self.aPsiteAnnot.shape[0],
                len(list(filter(lambda p:
                                    self.ncbi_tax_id in \
                                        self.dOrgSpecUniprot[(p[0], p[3], p[4])],
                                    self.aPsiteAnnot))),
                len(self.setSeqMismatchOrg),
                len(self.setSeqMismatchNameOrg)
            ))
    
    def collect_proteins(self):
        """
        Compiles a list of all kinases and substrates.
        """
        self.lAllProteins = set(self.dKinAssaySub.keys()) | \
            set(
                list(
                    itertools.chain(
                        *map(
                            lambda kss:
                                map(
                                    lambda ks:
                                        ks.domain.protein,
                                    kss
                                ),
                            self.dKinAssaySub.values()
                        )
                    )
                )
            )
    
    def load_network(self, pfile = None):
        if not pfile:
            self.pa.load_omnipath()
        else:
            self.pa.init_network(pfile = pfile)
        self.pa.get_directed()
        del self.pa.dgraph.es['ptm']
    
    def get_network(self, keep_also = [], step1 = True,
                    more_steps = None):
        """
        Creates a subnetwork based on certain criteria.
        
        :param list keep_also: List of additional vertex names
                               to include in core.
        :param bool step1: Whether include not only direct
                           links but one step indirect connections.
        :param int more_steps: Include longer indirect connections. This
                               param defines the number of mediator nodes.
        """
        if not hasattr(self, 'whole') or self.whole is None:
            self.whole = self.pa.dgraph
        
        new = copy.deepcopy(self.whole)
        self.pa.graph = new
        self.pa.dgraph = new
        self.pa.update_vname()
        self.pa.update_vindex()
        vids = dict(zip(new.vs['name'], range(new.vcount())))
        keep = list(map(lambda u: vids[u],
                    filter(lambda u: u in vids, self.lAllProteins)))
        keep_also = list(map(lambda u: vids[u],
                         filter(lambda u: u in vids, keep_also)))
        all_to_keep = set(keep) | set(keep_also)
        delete = set(range(new.vcount())) - all_to_keep
        
        keep_step1 = set([])
        for vid in delete:
            if len(set(self.pa.affected_by(vid)._vs) & all_to_keep):
                if len(set(self.pa._affected_by(vid)._vs) & all_to_keep):
                    keep_step1.add(vid)
        
        all_to_keep = all_to_keep | keep_step1
        delete = delete - keep_step1
        
        if more_steps is not None:
            
            n1 = self.pa._neighborhood(all_to_keep, order = more_steps, mode = 'IN')
            n2 = self.pa._neighborhood(all_to_keep, order = more_steps, mode = 'OUT')
            
            longer_paths = set(list(n1)) & set(list(n2))
            
            all_to_keep = all_to_keep | longer_paths
            delete = delete - longer_paths
        
        self.pa.graph.delete_vertices(list(delete))
        self.pa._directed = None
        self.pa.dgraph = self.pa.graph
        self.pa.update_vname()
        self.pa.update_vindex()
        
        sys.stdout.write('\t:: Nodes: %u -> %u, edges: %u -> %u\n'\
            '\t   Removed %u vertices and %u edges\n' % (
                self.whole.vcount(),
                self.pa.graph.vcount(),
                self.whole.ecount(),
                self.pa.graph.ecount(),
                len(delete),
                self.whole.ecount() - self.pa.graph.ecount()
            ))
    
    def sparsen_network(self, perc):
        """
        Removes a portion of the edges based on certain conditions.
        
        """
        
        dens0 = self.pa.graph.density()
        
        if 'ptm' not in self.pa.graph.es.attributes() or \
            max(map(len, self.pa.graph.es['ptm'])) == 0:
            self.pa.load_ptms()
        
        assay_ptms = set(itertools.chain(*self.dKinAssaySub.values()))
        
        # edges between kinases and substrates in the assay
        es_protected = list(filter(lambda e:
                                        len(set(e['ptm']) & assay_ptms),
                                    self.pa.graph.es))
        
        self.pa.graph.es['refxsrc'] = list(
            np.array(list(map(len, self.pa.graph.es['references']))) * \
            np.array(list(map(len, self.pa.graph.es['sources']))))
        
        self.pa.graph.es['sign'] = list(map(lambda e:
                                        e['dirs'].is_stimulation() or \
                                        e['dirs'].is_inhibition(),
                                    self.pa.graph.es))
        
        self.pa.graph.es['nptm'] = list(map(len, self.pa.graph.es['ptm']))
        
        self.pa.graph.es['mindeg'] = list(map(lambda e: min(
                                            self.pa.graph.vs[e.source].degree(),
                                            self.pa.graph.vs[e.target].degree()),
                                            self.pa.graph.es))
        
        self.pa.graph.es['weight'] = list(
            (np.array(self.pa.graph.es['refxsrc']) + 1) * \
            (np.array(self.pa.graph.es['sign'], dtype = np.int) + 1) * \
            (np.array(self.pa.graph.es['nptm']) + 1) / \
            np.array(self.pa.graph.es['mindeg']))
        
        cutoff = np.percentile(np.array(self.pa.graph.es['weight']), perc)
        
        delete = set(list(map(lambda e: e.index,
                        filter(lambda e:
                            e.index in es_protected or \
                            e['weight'] < cutoff,
                        self.pa.graph.es))))
        
        for v in self.pa.graph.vs:
            this_es = list(itertools.chain(
                self.pa.graph.es.select(_source = v.index),
                self.pa.graph.es.select(_target = v.index)
            ))
            
            eids = set(list(map(lambda e: e.index, this_es)))
            
            # if all edges of one node would be removed,
            # we still keep the one with the highest weight
            if len(eids - delete) == 0:
                
                maxweight = max(map(lambda e:
                                        (e['weight'], e.index),
                                    this_es),
                                key = lambda i: i[0])[1]
                
                if maxweight in delete:
                    delete.remove(maxweight)
        
        self.pa.graph.delete_edges(delete)
        self.pa.update_vindex()
        self.pa.update_vname()
        
        sys.stdout.write('\t:: Sparsening network: %u edges'\
            ' removed, %u have been kept.\n'\
            '\t   Density changed from %.04f to %.04f.\n' % \
                (len(delete), self.pa.graph.ecount(),
                 dens0, self.pa.graph.density()))